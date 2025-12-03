from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, distinct
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.database import get_db, SearchResult, SearchLink, init_db
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/excel/daily")
def export_daily_excel(
    days: int = 30,
    site_id: str = Query("default", description="Site ID"),
    db: Session = Depends(get_db)
):
    """Günlük link pozisyonlarını Excel olarak export eder"""
    since_date = datetime.utcnow() - timedelta(days=days)
    
    # Tüm arama sonuçlarını ve linklerini al
    results = db.query(SearchResult)\
        .filter(SearchResult.search_date >= since_date)\
        .order_by(SearchResult.search_date.asc())\
        .all()
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Günlük Pozisyonlar"
    
    # Header stilleri
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Header satırı
    headers = ["Tarih", "Saat", "URL", "Domain", "Başlık", "Pozisyon", "Snippet"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Veri satırları
    row = 2
    for result in results:
        links = db.query(SearchLink)\
            .filter(SearchLink.search_result_id == result.id)\
            .order_by(SearchLink.position.asc())\
            .all()
        
        for link in links:
            search_date = result.search_date
            ws.cell(row=row, column=1, value=search_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=search_date.strftime("%H:%M:%S"))
            ws.cell(row=row, column=3, value=link.url)
            ws.cell(row=row, column=4, value=link.domain or "")
            ws.cell(row=row, column=5, value=link.title or "")
            ws.cell(row=row, column=6, value=link.position)
            ws.cell(row=row, column=7, value=link.snippet or "")
            row += 1
    
    # Kolon genişliklerini ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 60
    
    # Excel dosyasını memory'de oluştur
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"google_search_bot_daily_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/position-history")
def export_position_history(
    url: str = None,
    days: int = 30,
    site_id: str = Query("default", description="Site ID"),
    db: Session = Depends(get_db)
):
    """Belirli bir URL'in pozisyon geçmişini Excel olarak export eder"""
    since_date = datetime.utcnow() - timedelta(days=days)
    
    query = db.query(SearchLink, SearchResult)\
        .join(SearchResult, SearchLink.search_result_id == SearchResult.id)\
        .filter(SearchResult.search_date >= since_date)
    
    if url:
        query = query.filter(SearchLink.url == url)
    
    results = query.order_by(SearchResult.search_date.asc()).all()
    
    if not results:
        raise HTTPException(status_code=404, detail="Veri bulunamadı")
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Pozisyon Geçmişi"
    
    # Header stilleri
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Header satırı
    headers = ["Tarih", "Saat", "URL", "Domain", "Başlık", "Pozisyon", "Değişim"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Veri satırları
    row = 2
    prev_position = None
    for link, result in results:
        search_date = result.search_date
        change = ""
        if prev_position is not None:
            diff = prev_position - link.position
            if diff > 0:
                change = f"↑ {diff}"
            elif diff < 0:
                change = f"↓ {abs(diff)}"
            else:
                change = "→"
        
        ws.cell(row=row, column=1, value=search_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=search_date.strftime("%H:%M:%S"))
        ws.cell(row=row, column=3, value=link.url)
        ws.cell(row=row, column=4, value=link.domain or "")
        ws.cell(row=row, column=5, value=link.title or "")
        ws.cell(row=row, column=6, value=link.position)
        ws.cell(row=row, column=7, value=change)
        
        # Pozisyon değişimine göre renklendir
        if change.startswith("↑"):
            ws.cell(row=row, column=7).fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        elif change.startswith("↓"):
            ws.cell(row=row, column=7).fill = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
        
        prev_position = link.position
        row += 1
    
    # Kolon genişliklerini ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    
    # Excel dosyasını memory'de oluştur
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"position_history_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/summary")
def export_summary_excel(
    days: int = 30,
    site_id: str = Query("default", description="Site ID"),
    db: Session = Depends(get_db)
):
    """Özet istatistikleri Excel olarak export eder"""
    since_date = datetime.utcnow() - timedelta(days=days)
    
    # Link istatistiklerini hesapla
    stats = db.query(
        SearchLink.url,
        SearchLink.domain,
        SearchLink.title,
        func.count(SearchLink.id).label("total_appearances"),
        func.min(SearchLink.created_at).label("first_seen"),
        func.max(SearchLink.created_at).label("last_seen"),
        func.avg(SearchLink.position).label("avg_position"),
        func.min(SearchLink.position).label("best_position"),
        func.max(SearchLink.position).label("worst_position")
    ).join(
        SearchResult, SearchLink.search_result_id == SearchResult.id
    ).filter(
        SearchResult.search_date >= since_date
    ).group_by(
        SearchLink.url, SearchLink.domain, SearchLink.title
    ).order_by(
        func.count(SearchLink.id).desc()
    ).all()
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Özet İstatistikler"
    
    # Header stilleri
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Header satırı
    headers = ["URL", "Domain", "Başlık", "Toplam Görünme", "İlk Görünme", "Son Görünme", 
               "Ort. Pozisyon", "En İyi", "En Kötü", "Aktif Gün"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Veri satırları
    row = 2
    for stat in stats:
        days_active = (stat.last_seen - stat.first_seen).days + 1
        
        ws.cell(row=row, column=1, value=stat.url)
        ws.cell(row=row, column=2, value=stat.domain or "")
        ws.cell(row=row, column=3, value=stat.title or "")
        ws.cell(row=row, column=4, value=stat.total_appearances)
        ws.cell(row=row, column=5, value=stat.first_seen.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=6, value=stat.last_seen.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=7, value=round(float(stat.avg_position), 2))
        ws.cell(row=row, column=8, value=stat.best_position)
        ws.cell(row=row, column=9, value=stat.worst_position)
        ws.cell(row=row, column=10, value=days_active)
        row += 1
    
    # Kolon genişliklerini ayarla
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    
    # Excel dosyasını memory'de oluştur
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pdf/daily")
def export_daily_pdf(
    days: int = 30,
    site_id: str = Query("default", description="Site ID"),
    db: Session = Depends(get_db)
):
    """Günlük link pozisyonlarını PDF olarak export eder"""
    init_db(site_id)
    since_date = datetime.utcnow() - timedelta(days=days)
    
    # Tüm arama sonuçlarını ve linklerini al
    results = db.query(SearchResult)\
        .filter(SearchResult.search_date >= since_date)\
        .order_by(SearchResult.search_date.asc())\
        .all()
    
    # PDF oluştur
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Başlık
    elements.append(Paragraph("Google Search Bot - Günlük Rapor", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tarih aralığı
    date_range = Paragraph(
        f"<b>Tarih Aralığı:</b> {(datetime.utcnow() - timedelta(days=days)).strftime('%d.%m.%Y')} - {datetime.utcnow().strftime('%d.%m.%Y')}",
        styles['Normal']
    )
    elements.append(date_range)
    elements.append(Spacer(1, 0.3*inch))
    
    # Veri tablosu
    data = [["Tarih", "Saat", "URL", "Domain", "Pozisyon"]]
    
    for result in results:
        links = db.query(SearchLink)\
            .filter(SearchLink.search_result_id == result.id)\
            .order_by(SearchLink.position.asc())\
            .all()
        
        for link in links:
            search_date = result.search_date
            # URL'yi kısalt (çok uzunsa)
            url_display = link.url[:60] + "..." if len(link.url) > 60 else link.url
            data.append([
                search_date.strftime("%Y-%m-%d"),
                search_date.strftime("%H:%M"),
                url_display,
                link.domain or "",
                str(link.position)
            ])
    
    # Tablo oluştur
    table = Table(data, colWidths=[1*inch, 0.8*inch, 3*inch, 1.2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # PDF'i oluştur
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"google_search_bot_daily_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pdf/summary")
def export_summary_pdf(
    days: int = 30,
    site_id: str = Query("default", description="Site ID"),
    db: Session = Depends(get_db)
):
    """Özet istatistikleri PDF olarak export eder"""
    init_db(site_id)
    since_date = datetime.utcnow() - timedelta(days=days)
    
    # Link istatistiklerini hesapla
    stats = db.query(
        SearchLink.url,
        SearchLink.domain,
        SearchLink.title,
        func.count(SearchLink.id).label("total_appearances"),
        func.min(SearchLink.created_at).label("first_seen"),
        func.max(SearchLink.created_at).label("last_seen"),
        func.avg(SearchLink.position).label("avg_position"),
        func.min(SearchLink.position).label("best_position"),
        func.max(SearchLink.position).label("worst_position")
    ).join(
        SearchResult, SearchLink.search_result_id == SearchResult.id
    ).filter(
        SearchResult.search_date >= since_date
    ).group_by(
        SearchLink.url, SearchLink.domain, SearchLink.title
    ).order_by(
        func.count(SearchLink.id).desc()
    ).limit(50).all()
    
    # PDF oluştur
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1
    )
    
    # Başlık
    elements.append(Paragraph("Google Search Bot - Özet Rapor", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tarih aralığı
    date_range = Paragraph(
        f"<b>Tarih Aralığı:</b> {(datetime.utcnow() - timedelta(days=days)).strftime('%d.%m.%Y')} - {datetime.utcnow().strftime('%d.%m.%Y')}",
        styles['Normal']
    )
    elements.append(date_range)
    elements.append(Spacer(1, 0.3*inch))
    
    # Veri tablosu
    data = [["URL", "Domain", "Görünme", "Ort. Pozisyon", "En İyi", "En Kötü"]]
    
    for stat in stats:
        url_display = stat.url[:40] + "..." if len(stat.url) > 40 else stat.url
        data.append([
            url_display,
            stat.domain or "",
            str(stat.total_appearances),
            f"{float(stat.avg_position):.1f}",
            str(stat.best_position),
            str(stat.worst_position)
        ])
    
    # Tablo oluştur
    table = Table(data, colWidths=[2.5*inch, 1.5*inch, 0.8*inch, 1*inch, 0.7*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # PDF'i oluştur
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"google_search_bot_summary_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




